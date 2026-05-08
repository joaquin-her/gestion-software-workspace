import argparse
import openpyxl
import re
import json
import os

def step1_extract_and_reformat():
    # 1. extract_backlog_us
    print("Extrayendo hoja 'Backlog-US' desde Artefactos.xlsx...")
    wb_source = openpyxl.load_workbook('artefactos/Artefactos.xlsx', data_only=True)
    if 'Backlog-US' not in wb_source.sheetnames:
        raise ValueError("No se encontró la hoja 'Backlog-US' en Artefactos.xlsx")
    
    wb_dest = openpyxl.Workbook()
    wb_dest.remove(wb_dest.active)
    
    ws_dest = wb_dest.create_sheet('Backlog-US')
    ws_source = wb_source['Backlog-US']
    
    for row in ws_source.iter_rows(values_only=True):
        ws_dest.append(row)
    
    os.makedirs('artefactos/output/backlog', exist_ok=True)
    out_file = 'artefactos/output/backlog/Backlog-US.xlsx'
    
    # 2. reformat
    print("Reformateando la hoja 'Backlog-US'...")
    if 'Backlog-reformated' in wb_dest.sheetnames:
        wb_dest.remove(wb_dest['Backlog-reformated'])
    
    ws_new = wb_dest.create_sheet('Backlog-reformated')
    ws_new.append(["Numero", "Titulo", "Prioridad", "Estimacion", "Descripcion", "Criterios de aceptacion"])
    
    def clean_value(prefix, val):
        if not val or val == 'None':
            return None
        val = str(val).strip()
        pattern = re.compile(f'^{prefix}\\s*:?\\s*', re.IGNORECASE)
        val = pattern.sub('', val)
        if val == '':
            return None
        return val

    rows = list(ws_dest.iter_rows(values_only=True))
    for i in range(0, len(rows), 4):
        row1 = rows[i]
        if not row1 or all(x is None for x in row1):
            continue
        row2 = rows[i+1] if i+1 < len(rows) else []
        row3 = rows[i+2] if i+2 < len(rows) else []
        
        nro = clean_value('Nro', row1[0] if len(row1) > 0 else None)
        titulo = clean_value('T.tulo', row1[1] if len(row1) > 1 else None)
        prioridad = clean_value('Prioridad', row1[2] if len(row1) > 2 else None)
        estimacion = clean_value('Estimaci.n', row1[3] if len(row1) > 3 else None)
        desc = clean_value('Descripci.n', row2[0] if len(row2) > 0 else None)
        crit = clean_value('Criterios de Aceptaci.n', row3[0] if len(row3) > 0 else None)
        
        ws_new.append([nro, titulo, prioridad, estimacion, desc, crit])

    wb_dest.save(out_file)
    print(f"-> Excel guardado en {out_file} con la hoja 'Backlog-reformated'.")

    # 3. extract_to_json
    print("Generando JSON desde 'Backlog-reformated'...")
    tickets = []
    rows = list(ws_new.iter_rows(values_only=True))
    headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(rows[0])]

    for row in rows[1:]:
        if not any(row):
            continue
        ticket = {}
        for i, col_name in enumerate(headers):
            val = row[i] if i < len(row) else None
            ticket[col_name] = None if (val == 'None' or val is None) else str(val).strip()
        tickets.append(ticket)

    os.makedirs('data', exist_ok=True)
    with open('data/backlog.json', 'w', encoding='utf-8') as f:
        json.dump(tickets, f, indent=2, ensure_ascii=False)
    
    print(f"-> Exportado JSON con {len(tickets)} historias a data/backlog.json")

def step2_generate_jira_tickets(backlog_file='data/backlog.json', tareas_file='data/tareas.json', output_file='data/jira_tickets.json'):
    print(f"Generando tickets formato Jira desde {backlog_file} y {tareas_file}...")
    if not os.path.exists(backlog_file) or not os.path.exists(tareas_file):
        print(f"Error: No se encuentran los archivos {backlog_file} o {tareas_file}")
        return

    with open(backlog_file, 'r', encoding='utf-8') as f:
        backlog = json.load(f)

    with open(tareas_file, 'r', encoding='utf-8') as f:
        tareas = json.load(f)

    jira_tickets = []
    issue_id_counter = 1

    for story in backlog:
        if not story.get('Titulo'):
            continue
            
        story_id = str(issue_id_counter)
        issue_id_counter += 1
        
        desc = story.get('Descripcion', '')
        if story.get('Criterios de aceptacion'):
            desc += '\n\n*Criterios de Aceptación:*\n' + story['Criterios de aceptacion']
        
        # Si la historia viene del USM, puede tener metadatos extra
        if story.get('Rol'):
            desc = f"*Rol:* {story['Rol']}\n*Categoría:* {story.get('Categoria', 'N/A')}\n\n" + desc

        story_points = None
        if story.get('Estimacion') and str(story['Estimacion']).isdigit():
            story_points = int(story['Estimacion'])

        # Preservar etiquetas si existen (caso USM)
        story_labels = story.get('Labels', [])

        jira_tickets.append({
            "summary": story['Titulo'],
            "description": desc,
            "issueType": "Story",
            "storyPoints": story_points,
            "issueId": story_id,
            "labels": story_labels
        })
        
        historia_tareas = next((t for t in tareas if t['Historia'] == story['Titulo']), None)
        if historia_tareas and 'Tareas' in historia_tareas:
            for tarea in historia_tareas['Tareas']:
                task_id = str(issue_id_counter)
                issue_id_counter += 1
                
                # Las subtareas heredan etiquetas de la historia + su propio tipo
                task_labels = story_labels + [tarea['Tipo'].replace(" ", "")]
                
                jira_tickets.append({
                    "summary": tarea['Titulo'],
                    "description": f"Sub-tarea del área {tarea['Tipo']} para la historia: {story['Titulo']}",
                    "issueType": "Sub-task",
                    "parentId": story_id,
                    "labels": task_labels
                })

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(jira_tickets, f, indent=2, ensure_ascii=False)
    print(f"-> Generado {output_file} con {len(jira_tickets)} tickets.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Pipeline del Jira Backlog Importer")
    parser.add_argument('--pre-ai', action='store_true', help='(Paso 1) Extrae y formatea el Excel, y genera data/backlog.json')
    parser.add_argument('--post-ai', action='store_true', help='(Paso 3) Empaqueta data/backlog.json y data/tareas.json en formato Jira')
    parser.add_argument('--usm', action='store_true', help='Indica que se debe procesar el archivo del USM')
    args = parser.parse_args()

    if args.pre_ai:
        step1_extract_and_reformat()
    elif args.post_ai:
        if args.usm:
            step2_generate_jira_tickets(
                backlog_file='data/usm_backlog.json',
                tareas_file='data/tareas_usm.json',
                output_file='data/jira_tickets_usm.json'
            )
        else:
            step2_generate_jira_tickets()
    else:
        parser.print_help()
