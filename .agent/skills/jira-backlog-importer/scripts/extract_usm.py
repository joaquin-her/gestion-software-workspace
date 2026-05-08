import openpyxl
import json
import os
import re

def clean_label(text):
    """Limpia el texto para usarlo como etiqueta de Jira (sin espacios, alfanumérico)"""
    # Remueve acentos y caracteres especiales básicos
    import unicodedata
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
    # Remueve todo lo que no sea letra o número
    text = re.sub(r'[^a-zA-Z0-9\s]', '', text)
    # Convierte a PascalCase (TitleCase sin espacios)
    return "".join(word.capitalize() for word in text.split())

def extract_section(ws, start_row, end_row, role_name):
    # La fila de categoría es la segunda fila del rango indicado
    category_row = start_row + 1
    
    categories = {}
    current_category = "General"
    
    # Recorremos la fila de categorías de izquierda a derecha.
    # Si hay celdas combinadas o vacías, arrastramos el valor de la categoría anterior.
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=category_row, column=col).value
        if cell_val and str(cell_val).strip():
            current_category = str(cell_val).strip()
        categories[col] = current_category
        
    stories = []
    # Las historias están desde la fila debajo de las categorías hasta el final del rango
    for r in range(category_row + 1, end_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip():
                historia = str(val).strip()
                categoria = categories[c]
                
                # Razonamiento de etiquetas:
                # 1. El Rol como etiqueta ayuda a filtrar para quién es la feature.
                # 2. La Categoría (ej: Autentificación, Agenda) ayuda a agrupar épicas o módulos.
                labels = [
                    clean_label(role_name),
                    clean_label(categoria)
                ]
                
                stories.append({
                    "Titulo": historia,
                    "Rol": role_name,
                    "Categoria": categoria,
                    "Labels": labels,
                    "FilaOrigen": r,
                    "ColumnaOrigen": c
                })
    return stories

def main():
    file_path = 'artefactos/Artefactos.xlsx'
    print(f"Leyendo archivo: {file_path}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if 'USM-Corregido' not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'USM-Corregido'")
        
    ws = wb['USM-Corregido']
    all_stories = []
    
    # Profesionales independientes: Fila 1 a 9
    print("Extrayendo sección: Profesionales Independientes (filas 1-9)")
    all_stories.extend(extract_section(ws, 1, 9, "Profesional independiente"))

    # Cliente-Paciente: Fila 13 a 20
    print("Extrayendo sección: Cliente-Paciente (filas 13-20)")
    all_stories.extend(extract_section(ws, 13, 20, "Cliente Paciente"))

    # Grupo de profesionales: Fila 22 a 27
    print("Extrayendo sección: Grupo de profesionales (filas 22-27)")
    all_stories.extend(extract_section(ws, 22, 27, "Grupo de profesionales"))

    os.makedirs('data', exist_ok=True)
    out_path = 'data/usm_backlog.json'
    
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(all_stories, f, indent=2, ensure_ascii=False)

    print(f"\n¡Éxito! Se extrajeron {len(all_stories)} historias de usuario del USM.")
    print(f"El resultado con roles y categorías ha sido guardado en: {out_path}")

if __name__ == "__main__":
    main()
